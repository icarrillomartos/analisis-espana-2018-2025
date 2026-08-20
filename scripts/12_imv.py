# -*- coding: utf-8 -*-
"""Ingreso Mínimo Vital (IMV): nómina mensual y agregado anual.

Fuente: Seguridad Social (INSS), estadística "Ingreso Mínimo Vital · Nóminas".
El portal publica un XLSX por mes con el detalle por comunidad autónoma y provincia;
aquí se toma únicamente la fila TOTAL nacional de cada fichero.

Aviso de cobertura: la estadística mensual del IMV solo está publicada como fichero
descargable desde enero de 2024, pese a que la prestación existe desde junio de 2020.
La única cifra que alcanza 2020 es la columna "Importe bruto acumulado", que suma
desde el inicio de la prestación.

Salidas:
  data_external/imv_raw/*.xlsx          (ficheros originales, sin modificar)
  data_external/imv_nomina_mensual.csv
  data_external/imv_anual.csv
  filas añadidas a data_external/metadata.csv
"""
import os, re, csv, html, unicodedata, urllib.request, urllib.parse

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data_external")
CRUDO = os.path.join(OUT, "imv_raw")
HOY = "2026-08-20"

HOST = "https://www.seg-social.es"
BASE = (HOST + "/wps/portal/wss/internet/EstadisticasPresupuestosEstudios/Estadisticas"
        "/cbe2fda1-3ac7-4bc8-a5ec-06c178839e11/d1fee385-f3ad-4e58-a129-d8818d93ea4b")
# Páginas índice de "Nóminas" por año. Son los únicos años publicados por el organismo.
INDICES = {2024: "ae6d943c-a7eb-4112-aa8e-002f268f5d73",
           2025: "531c050e-e879-4de4-bff7-099af492edd7",
           2026: "ef51a713-7a93-4a5e-a6bb-9e5adf310de6"}

UA = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"}

MESES = {1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
         7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre",
         12: "diciembre"}

# Cabeceras del XLSX -> columna de salida. Se localizan por texto, no por posición.
CAMPOS = [("prestaciones",              r"numero de prestaciones"),
          ("beneficiarios",             r"numero\s+de beneficiarios"),
          ("importe_bruto_mes_eur",     r"importe bruto de las nominas"),
          ("cuantia_media_hogar_eur",   r"cuantia media mensual por hogar"),
          ("importe_bruto_acumulado_eur", r"importe bruto acumulado")]


def bajar(url, destino):
    if os.path.exists(destino) and os.path.getsize(destino) > 5000:
        return destino
    with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=120) as r:
        datos = r.read()
    with open(destino, "wb") as fh:
        fh.write(datos)
    return destino


def enlaces_nomina(anyo, uuid):
    """Rastrea la página índice del año y devuelve [(aaaamm, url_xlsx)]."""
    url = f"{BASE}/{uuid}"
    with urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=120) as r:
        pagina = r.read().decode("utf-8", "replace")
    vistos = {}
    for m in re.finditer(r'href="(/wps/wcm/connect/wss/[^"]+?\.xlsx[^"]*)"', pagina, re.I):
        href = html.unescape(m.group(1))
        nombre = urllib.parse.unquote(href.split("?")[0].split("/")[-1])
        if "mina" not in nombre.lower():          # solo "Nómina", no "Altas acumuladas"
            continue
        per = re.match(r"(\d{4})(\d{2})", nombre)
        if not per or int(per.group(1)) != anyo:
            continue
        # basta con MOD=AJPERES; se descarta el CACHEID, que es volátil
        vistos[per.group(0)] = HOST + href.split("?")[0] + "?MOD=AJPERES"
    return sorted(vistos.items())


def sin_tildes(s):
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def leer_total(ruta):
    """Devuelve el dict de la fila TOTAL nacional del XLSX de nómina."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    columnas = {}
    fila_total = None
    for i, fila in enumerate(filas):
        primera = sin_tildes(fila[0]) if fila and fila[0] is not None else ""
        if fila_total is None and primera == "total":
            fila_total = i
        if fila_total is not None:
            continue
        for j, celda in enumerate(fila):
            if celda is None:
                continue
            texto = sin_tildes(celda)
            for clave, patron in CAMPOS:
                if clave not in columnas and re.search(patron, texto):
                    columnas[clave] = j
    if fila_total is None:
        raise ValueError(f"{os.path.basename(ruta)}: no se encontró la fila TOTAL")

    fila = filas[fila_total]
    reg = {}
    for clave, _ in CAMPOS:
        j = columnas.get(clave)
        valor = fila[j] if (j is not None and j < len(fila)) else None
        reg[clave] = valor if isinstance(valor, (int, float)) else None
    return reg


os.makedirs(CRUDO, exist_ok=True)

mensual = []
urls_usadas = []
for anyo, uuid in sorted(INDICES.items()):
    for periodo, url in enlaces_nomina(anyo, uuid):
        destino = os.path.join(CRUDO, f"imv_nomina_{periodo}.xlsx")
        bajar(url, destino)
        reg = leer_total(destino)
        reg.update(anyo=int(periodo[:4]), mes=int(periodo[4:]),
                   periodo=f"{periodo[:4]}-{periodo[4:]}",
                   mes_nombre=MESES[int(periodo[4:])],
                   fichero_origen=os.path.basename(destino), url_origen=url)
        mensual.append(reg)
        urls_usadas.append(url)
        print(f"  {reg['periodo']}  prestaciones={reg['prestaciones']!s:>9}  "
              f"beneficiarios={reg['beneficiarios']!s:>9}  "
              f"importe={reg['importe_bruto_mes_eur']!s:>20}")

mensual.sort(key=lambda r: (r["anyo"], r["mes"]))
if not mensual:
    raise SystemExit("No se descargó ninguna nómina: revisar las páginas índice.")

COLS_MES = ["periodo", "anyo", "mes", "mes_nombre", "prestaciones", "beneficiarios",
            "importe_bruto_mes_eur", "cuantia_media_hogar_eur",
            "importe_bruto_acumulado_eur", "fichero_origen", "url_origen"]
ruta_mes = os.path.join(OUT, "imv_nomina_mensual.csv")
with open(ruta_mes, "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=COLS_MES, extrasaction="ignore")
    w.writeheader()
    w.writerows(mensual)

# ---- agregado anual -------------------------------------------------------
def media(vals):
    vals = [v for v in vals if v is not None]
    return round(sum(vals) / len(vals), 2) if vals else None

def de_diciembre(filas, campo):
    for r in filas:
        if r["mes"] == 12:
            return r[campo]
    return None

anual = []
for anyo in sorted({r["anyo"] for r in mensual}):
    filas = [r for r in mensual if r["anyo"] == anyo]
    importes = [r["importe_bruto_mes_eur"] for r in filas if r["importe_bruto_mes_eur"] is not None]
    completo = len(filas) == 12 and len(importes) == 12
    faltan = [MESES[m] for m in range(1, 13) if m not in {r["mes"] for r in filas}]
    anual.append({
        "anyo": anyo,
        "meses_disponibles": len(filas),
        "anyo_completo": "si" if completo else "no",
        "importe_bruto_anual_eur": round(sum(importes), 2) if importes else None,
        "prestaciones_media": media([r["prestaciones"] for r in filas]),
        "prestaciones_diciembre": de_diciembre(filas, "prestaciones"),
        "beneficiarios_media": media([r["beneficiarios"] for r in filas]),
        "beneficiarios_diciembre": de_diciembre(filas, "beneficiarios"),
        "cuantia_media_hogar_media": media([r["cuantia_media_hogar_eur"] for r in filas]),
        "importe_bruto_acumulado_dic_eur": de_diciembre(filas, "importe_bruto_acumulado_eur"),
        "aviso": "" if completo else
                 ("Año incompleto: la suma anual solo cubre los meses publicados. "
                  "Sin fichero XLSX para " + ", ".join(faltan) + "."),
    })

COLS_ANY = list(anual[0].keys())
ruta_anual = os.path.join(OUT, "imv_anual.csv")
with open(ruta_anual, "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=COLS_ANY)
    w.writeheader()
    w.writerows(anual)

# ---- metadatos ------------------------------------------------------------
ORG = "Seguridad Social (Ministerio de Inclusión, Seguridad Social y Migraciones)"
COBERTURA = f"{mensual[0]['periodo']} a {mensual[-1]['periodo']}"
NOTA_BASE = ("La serie descargable empieza en enero de 2024: la estadística mensual del IMV "
             "solo se publica como fichero desde entonces, aunque la prestación existe desde "
             "junio de 2020. No hay ficheros oficiales vivos para 2020-2023 (comprobado: los "
             "enlaces anteriores devuelven 404 y datos.gob.es no tiene ningún conjunto de IMV). "
             "La columna 'importe bruto acumulado' es la única referencia que llega hasta 2020, "
             "porque acumula desde el inicio de la prestación.")

meta = [
    {"fichero": "imv_nomina_mensual.csv",
     "indicador": "Ingreso Mínimo Vital: prestaciones (hogares), beneficiarios, importe bruto "
                  "de la nómina y cuantía media mensual por hogar",
     "organismo": ORG,
     "dataset": "Estadística del Ingreso Mínimo Vital · Nóminas (fila TOTAL nacional)",
     "url_exacta": mensual[-1]["url_origen"],
     "cobertura": COBERTURA,
     "geografia": "ES",
     "fecha_descarga": HOY,
     "notas": ("Un XLSX por mes; cada fichero tiene un UUID propio, por lo que el script "
               "rastrea las páginas índice anuales en lugar de construir las URL. La url_exacta "
               "es la del último mes descargado; el listado completo está en la columna "
               "url_origen del propio CSV. Julio de 2024 falta porque el organismo solo lo "
               "publicó en PDF, sin XLSX. " + NOTA_BASE)},
    {"fichero": "imv_anual.csv",
     "indicador": "Ingreso Mínimo Vital: gasto bruto anual y prestaciones/beneficiarios "
                  "(media del año y dato de diciembre)",
     "organismo": ORG,
     "dataset": "Estadística del Ingreso Mínimo Vital · Nóminas (agregado propio por año)",
     "url_exacta": f"{BASE}/{INDICES[max(INDICES)]}",
     "cobertura": f"{anual[0]['anyo']}-{anual[-1]['anyo']}",
     "geografia": "ES",
     "fecha_descarga": HOY,
     "notas": ("Agregado propio a partir de imv_nomina_mensual.csv: el importe anual es la suma "
               "de las nóminas mensuales, no una cifra publicada por el organismo. Los años sin "
               "12 meses quedan marcados con anyo_completo='no'. " + NOTA_BASE)},
]

mp = os.path.join(OUT, "metadata.csv")
prev = list(csv.DictReader(open(mp, encoding="utf-8"))) if os.path.exists(mp) else []
campos = list(prev[0].keys()) if prev else list(meta[0].keys())
prev = [p for p in prev if p["fichero"] not in {m["fichero"] for m in meta}]
with open(mp, "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=campos)
    w.writeheader()
    w.writerows(prev + meta)

print(f"\nimv_nomina_mensual.csv  {len(mensual)} meses  ({COBERTURA})")
print(f"imv_anual.csv           {len(anual)} años")
print(f"metadata.csv actualizado: {len(prev) + len(meta)} filas")
